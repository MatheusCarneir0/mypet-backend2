# apps/relatorios/services.py
"""
Services para geração de relatórios.
"""
from django.db.models import Count, Sum, Avg
from django.utils import timezone
from datetime import timedelta
from .models import Relatorio


def _get_total_pets() -> int:
    """Helper para evitar import circular ao contar pets."""
    from apps.pets.models import Pet
    return Pet.objects.filter(ativo=True).count()

class RelatorioService:
    """
    Service para geração de relatórios gerenciais.
    """
    
    @staticmethod
    def _get_report_data(tipo, filtros):
        from apps.agendamentos.models import Agendamento
        from apps.clientes.models import Cliente
        from apps.servicos.models import Servico
        from apps.historico.models import HistoricoAtendimento
        from django.utils.timezone import localtime
        
        data_inicio = filtros.get('data_inicio')
        data_fim = filtros.get('data_fim')
        
        def format_currency(val):
            val_str = f"{val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            return f"R$ {val_str}"
            
        def format_date(dt):
            if not dt: return ""
            return localtime(dt).strftime("%d/%m/%Y às %H:%M")
        
        headers = []
        rows = []
        
        if tipo == 'AGENDAMENTOS':
            headers = ['ID', 'Data/Hora', 'Cliente', 'Pet', 'Serviço', 'Status', 'Valor']
            qs = Agendamento.objects.all()
            if data_inicio: qs = qs.filter(data_hora__date__gte=data_inicio)
            if data_fim: qs = qs.filter(data_hora__date__lte=data_fim)
            total_faturamento = 0.0
            total_agendamentos = 0
            for obj in qs:
                val = float(obj.valor_pago) if obj.valor_pago else 0.0
                total_faturamento += val
                total_agendamentos += 1
                
                srv_nome = getattr(obj, 'servico_nome', None)
                if not srv_nome:
                    srv_nome = obj.servico.nome if getattr(obj, 'servico', None) else "Não informado"
                    
                rows.append([
                    str(obj.id), 
                    format_date(obj.data_hora), 
                    obj.cliente.usuario.nome if getattr(obj, 'cliente', None) and getattr(obj.cliente, 'usuario', None) else "", 
                    obj.pet.nome if getattr(obj, 'pet', None) else "", 
                    srv_nome, 
                    obj.get_status_display(), 
                    format_currency(val)
                ])
            
            # Linha de totais
            rows.append([f"TOTAL: {total_agendamentos} agendamentos  |  Faturamento: {format_currency(total_faturamento)}", "", "", "", "", "", ""])
                
        elif tipo == 'CLIENTES':
            headers = ['ID', 'Nome', 'CPF', 'Cidade', 'Estado']
            qs = Cliente.objects.all()
            if data_inicio: qs = qs.filter(data_criacao__date__gte=data_inicio)
            if data_fim: qs = qs.filter(data_criacao__date__lte=data_fim)
            for obj in qs:
                rows.append([str(obj.id), obj.usuario.nome if getattr(obj, 'usuario', None) else "", obj.cpf, obj.cidade, obj.estado])
                
        elif tipo == 'SERVICOS':
            headers = ['ID', 'Nome', 'Preço', 'Duração (min)']
            qs = Servico.objects.all()
            if data_inicio: qs = qs.filter(data_criacao__date__gte=data_inicio)
            if data_fim: qs = qs.filter(data_criacao__date__lte=data_fim)
            for obj in qs:
                val = float(obj.preco) if obj.preco else 0.0
                rows.append([str(obj.id), obj.nome, format_currency(val), str(obj.duracao_minutos)])
                
        elif tipo == 'FATURAMENTO':
            headers = ['ID', 'Data', 'Pet', 'Serviço', 'Valor Pago']
            qs = HistoricoAtendimento.objects.all()
            if data_inicio: qs = qs.filter(data_atendimento__date__gte=data_inicio)
            if data_fim: qs = qs.filter(data_atendimento__date__lte=data_fim)
            for obj in qs:
                val = float(obj.valor_pago) if obj.valor_pago else 0.0
                rows.append([str(obj.id), format_date(obj.data_atendimento), obj.pet.nome if getattr(obj, 'pet', None) else "", obj.tipo_servico, format_currency(val)])
                
        return headers, rows

    @staticmethod
    def _generate_csv(relatorio, headers, rows):
        import csv
        import io
        from django.core.files.base import ContentFile
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        file_name = f"relatorio_{relatorio.id}.csv"
        relatorio.arquivo.save(file_name, ContentFile(output.getvalue().encode('utf-8')))

    @staticmethod
    def _generate_excel(relatorio, headers, rows):
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        import io
        from django.core.files.base import ContentFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatorio"
        ws.append(headers)
        for row in rows:
            ws.append(row)
            if str(row[0]).startswith("TOTAL:"):
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.font = Font(bold=True)
        output = io.BytesIO()
        wb.save(output)
        file_name = f"relatorio_{relatorio.id}.xlsx"
        relatorio.arquivo.save(file_name, ContentFile(output.getvalue()))

    @staticmethod
    def _generate_pdf(relatorio, headers, rows):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus.flowables import HRFlowable
        from django.utils.timezone import localtime
        import io
        from django.core.files.base import ContentFile
        from datetime import datetime
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output, 
            pagesize=landscape(A4), 
            rightMargin=2*cm, 
            leftMargin=2*cm, 
            topMargin=2*cm, 
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        azul_escuro = colors.HexColor('#1a3a5c')
        cinza_claro = colors.HexColor('#f1f5f9')
        cinza_medio = colors.HexColor('#e2e8f0')
        verde_escuro = colors.HexColor('#166534')
        vermelho_escuro = colors.HexColor('#991b1b')
        azul_status = colors.HexColor('#1e40af')
        
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, textColor=azul_escuro, spaceAfter=6)
        subtitle_style = ParagraphStyle('SubTitleStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.black, spaceAfter=12)
        info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.gray, spaceAfter=2)
        info_style_9pt = ParagraphStyle('Info9', parent=info_style, fontSize=9)
        
        elements.append(Paragraph("FarmaVet", title_style))
        elements.append(Paragraph(f"Relatório de {relatorio.get_tipo_display()}", subtitle_style))
        
        data_inicio = relatorio.filtros.get('data_inicio')
        data_fim = relatorio.filtros.get('data_fim')
        periodo_str = "Todo o período"
        
        try:
            di_str, df_str = "", ""
            if data_inicio: 
                d_i = datetime.fromisoformat(data_inicio)
                di_str = d_i.strftime("%d/%m/%Y")
            if data_fim:
                d_f = datetime.fromisoformat(data_fim)
                df_str = d_f.strftime("%d/%m/%Y")
            
            if di_str and df_str:
                periodo_str = f"{di_str} a {df_str}"
            elif di_str:
                periodo_str = f"A partir de {di_str}"
            elif df_str:
                periodo_str = f"Até {df_str}"
        except Exception:
            pass

        elements.append(Paragraph(f"Período: {periodo_str}", info_style))
        gerado_em = localtime(relatorio.data_geracao).strftime("%d/%m/%Y às %H:%M")
        elements.append(Paragraph(f"Gerado em: {gerado_em}", info_style_9pt))
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.gray, spaceAfter=20))
        
        table_data = [headers]
        status_col_idx = headers.index('Status') if 'Status' in headers else None
        
        ts = [
            ('BACKGROUND', (0,0), (-1,0), azul_escuro),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8.5),
        ]
        
        for i, row in enumerate(rows):
            is_total_row = str(row[0]).startswith("TOTAL:")
            if is_total_row:
                table_data.append([row[0]] + [''] * (len(headers) - 1))
            else:
                table_data.append(row)
            
            row_idx = i + 1 
            
            if is_total_row:
                ts.append(('BACKGROUND', (0, row_idx), (-1, row_idx), cinza_medio))
                ts.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                ts.append(('SPAN', (0, row_idx), (-1, row_idx)))
                ts.append(('ALIGN', (0, row_idx), (-1, row_idx), 'RIGHT'))
            else:
                bg_color = colors.white if i % 2 == 0 else cinza_claro
                ts.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))
                
                if status_col_idx is not None:
                    status_val = str(row[status_col_idx]).strip().lower()
                    if status_val == 'concluído':
                        ts.append(('TEXTCOLOR', (status_col_idx, row_idx), (status_col_idx, row_idx), verde_escuro))
                        ts.append(('FONTNAME', (status_col_idx, row_idx), (status_col_idx, row_idx), 'Helvetica-Bold'))
                    elif status_val == 'cancelado':
                        ts.append(('TEXTCOLOR', (status_col_idx, row_idx), (status_col_idx, row_idx), vermelho_escuro))
                        ts.append(('FONTNAME', (status_col_idx, row_idx), (status_col_idx, row_idx), 'Helvetica-Bold'))
                    elif status_val == 'agendado':
                        ts.append(('TEXTCOLOR', (status_col_idx, row_idx), (status_col_idx, row_idx), azul_status))
                        ts.append(('FONTNAME', (status_col_idx, row_idx), (status_col_idx, row_idx), 'Helvetica-Bold'))

        col_widths = None
        if relatorio.tipo == 'AGENDAMENTOS':
            # Landscape A4 = 29.7cm. Margins = 4cm. Available = 25.7cm
            col_widths = [1.5*cm, 3.5*cm, 6.0*cm, 3.7*cm, 5.5*cm, 3.0*cm, 2.5*cm]
        elif relatorio.tipo == 'CLIENTES':
            col_widths = [1.5*cm, 7.0*cm, 4.0*cm, 6.0*cm, 2.0*cm]
        elif relatorio.tipo == 'SERVICOS':
            col_widths = [1.5*cm, 10.0*cm, 4.0*cm, 5.0*cm]
        elif relatorio.tipo == 'FATURAMENTO':
            col_widths = [1.5*cm, 4.0*cm, 5.0*cm, 6.0*cm, 4.0*cm]

        if col_widths and len(col_widths) == len(headers):
            t = Table(table_data, colWidths=col_widths)
        else:
            t = Table(table_data)

        t.setStyle(TableStyle(ts))
        elements.append(t)
        
        def draw_footer(canvas, doc):
            canvas.saveState()
            canvas.setStrokeColor(colors.gray)
            canvas.setLineWidth(0.5)
            # Margen = 2cm, Landscape A4 Width = 29.7cm (841 points)
            canvas.line(2*cm, 1.5*cm, landscape(A4)[0]-2*cm, 1.5*cm)
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            canvas.drawString(2*cm, 1.0*cm, "FarmaVet - Sistema de Gestão para Pet Shop e Clínica Veterinária")
            canvas.drawRightString(landscape(A4)[0]-2*cm, 1.0*cm, f"Página {canvas.getPageNumber()}")
            canvas.restoreState()
            
        doc.build(elements, onFirstPage=draw_footer, onLaterPages=draw_footer)
        
        file_name = f"relatorio_{relatorio.id}.pdf"
        relatorio.arquivo.save(file_name, ContentFile(output.getvalue()))

    @staticmethod
    def gerar_relatorio(administrador, tipo, formato, filtros):
        """
        Gera um relatório baseado nos parâmetros fornecidos.
        
        Args:
            administrador: Usuário administrador solicitante
            tipo: Tipo do relatório
            formato: Formato de saída (PDF, Excel, CSV)
            filtros: Dicionário com filtros aplicados
        
        Returns:
            Relatorio: Instância do relatório gerado
        """
        import datetime
        
        filtros_formatados = {}
        if filtros:
            for k, v in filtros.items():
                if isinstance(v, (datetime.date, datetime.datetime)):
                    filtros_formatados[k] = v.isoformat()
                else:
                    filtros_formatados[k] = v

        # Criar registro do relatório
        relatorio = Relatorio.objects.create(
            administrador=administrador,
            tipo=tipo,
            formato=formato,
            filtros=filtros_formatados
        )
        
        # Gerar os dados
        headers, rows = RelatorioService._get_report_data(tipo, filtros_formatados)
        
        if formato == 'CSV':
            RelatorioService._generate_csv(relatorio, headers, rows)
        elif formato == 'EXCEL':
            RelatorioService._generate_excel(relatorio, headers, rows)
        elif formato == 'PDF':
            RelatorioService._generate_pdf(relatorio, headers, rows)
            
        relatorio.status = 'CONCLUIDO'
        relatorio.save()
        
        from django.core.mail import send_mail
        from django.conf import settings
        
        try:
            assunto = f"Relatório {relatorio.get_tipo_display()} gerado com sucesso"
            mensagem = (
                f"Seu relatório foi gerado em {relatorio.data_geracao.strftime('%d/%m/%Y %H:%M')}.\n"
                f"Acesse o sistema para fazer o download: "
                f"{settings.FRONTEND_URL}/admin/relatorios\n"
                f"Arquivo disponível em: {settings.BACKEND_URL}{relatorio.arquivo.url}"
            )
            send_mail(
                assunto,
                mensagem,
                settings.DEFAULT_FROM_EMAIL,
                [relatorio.administrador.email],
                fail_silently=True
            )
        except Exception:
            pass
        
        return relatorio
    
    @staticmethod
    def obter_dashboard_data():
        """
        Retorna dados para dashboard gerencial.
        """
        hoje = timezone.now().date()
        inicio_mes = hoje.replace(day=1)
        
        # Agendamentos do mês
        from apps.agendamentos.models import Agendamento
        agendamentos_mes = Agendamento.objects.filter(
            data_hora__gte=inicio_mes,
            ativo=True
        )
        
        # Faturamento do mês
        from apps.historico.models import HistoricoAtendimento
        faturamento_mes = HistoricoAtendimento.objects.filter(
            data_atendimento__gte=inicio_mes
        ).aggregate(
            total=Sum('valor_pago')
        )['total'] or 0
        
        # Novos clientes do mês
        from apps.clientes.models import Cliente
        novos_clientes_mes = Cliente.objects.filter(
            data_criacao__gte=inicio_mes,
            ativo=True
        ).count()
        
        # Serviços mais realizados
        servicos_top = HistoricoAtendimento.objects.filter(
            data_atendimento__gte=inicio_mes
        ).values('tipo_servico').annotate(
            quantidade=Count('id')
        ).order_by('-quantidade')[:5]
        
        return {
            'total_agendamentos_mes': agendamentos_mes.count(),
            'faturamento_mes': float(faturamento_mes),
            'novos_clientes_mes': novos_clientes_mes,
            'servicos_top': list(servicos_top),
            'agendamentos_hoje': agendamentos_mes.filter(
                data_hora__date=hoje
            ).count(),
            # #4: adicionado para evitar que o frontend faça requests extras
            'total_clientes': Cliente.objects.filter(ativo=True).count(),
            'total_pets': _get_total_pets(),
        }

